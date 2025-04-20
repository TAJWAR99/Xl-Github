import openpyxl
import os
import subprocess
from .createIssue import create_github_issue
from .getIssueUrl import getIssueUrl

try:
    from .google_sheets_config import EXCEL_SHEET
except ImportError:
    EXCEL_SHEET = None

def monitor_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, EXCEL_SHEET)

    if not os.path.exists(file_path):
        raise FileNotFoundError("File not found. Please ensure you have properly configured the google_sheets_config file or added the file in correct directory.")

    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:  # Iterate through all sheet names
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]  # Access each sheet

        # Get column headers from the first row
        headers = [cell.value for cell in sheet[1]]
        # Create a mapping from header names to column indices (1-based)
        header_index_map = {header: idx + 1 for idx, header in enumerate(headers)}

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
            row_dict = dict(zip(headers, row))  # Convert row to a dictionary
            
            test_case_id = row_dict.get("TestCase ID", "")
            test_case_name = row_dict.get("TestCase Name", "")
            status = row_dict.get("Status", "")
            issue_created = row_dict.get("Issue Status", "")
            assignee = row_dict.get("Assignee", "")
            project_name = row_dict.get("Project", "")
            module_name = row_dict.get("Module", "")
            sprint_name = row_dict.get("Sprint", "")
            
            if status.lower() == "failed" and issue_created.lower() != "yes":
                issue_title = f"{test_case_id}:{test_case_name}"
                issue_body = f"Test Case Name: {test_case_name}\nStatus: {status}"
                created, issue_url = create_github_issue(issue_title, issue_body, assignee, ["bug", "test-failure"], project_name, module_name, sprint_name)
                if created:
                    # Update the Excel file
                    sheet.cell(row=row_idx, column=header_index_map["Issue Status"]).value = "Yes"
                    sheet.cell(row=row_idx, column=header_index_map["Issue Url"]).value = getIssueUrl(issue_url)


    # Save the workbook with the added issue links
    workbook.save(file_path)
    # "touch" the file (update its timestamp) to force VS Code to detect a change
    os.utime(file_path, None)

    workbook.close()